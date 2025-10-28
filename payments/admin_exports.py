import csv
import json
from datetime import datetime
from io import BytesIO
from typing import Iterable

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Transaction, PurchasingItem

TRANSACTION_HEADERS = [
    "ID",
    "User ID",
    "First Name",
    "Last Name",
    "Email",
    "Phone Number",
    "National Code",
    "Gender",
    "City",
    "University",
    "Major",
    "Amount (IRR)",
    "Amount (Toman)",
    "Status",
    "Items",
    "Discount Code",
    "Track ID",
    "Order ID",
    "Reference Number",
    "Card Number",
    "Result",
    "Gateway Response",
    "Created At",
    "Completed At",
    "Updated At",
]

PURCHASING_ITEM_HEADERS = [
    "Full Name",
    "Email",
    "Phone Number",
    "National Code",
    "Gender",
    "City",
    "University",
    "Major",
    "Payment Reference",
    "Track ID",
    "Order ID",
    "Status",
    "Amount (Toman)",
    "Completed At",
]


def _flatten_items(transaction: Transaction) -> str:
    try:
        items = json.loads(transaction.items or "[]")
        if isinstance(items, list):
            return ", ".join(items)
        return str(items)
    except (TypeError, json.JSONDecodeError):
        return transaction.items or ""



def _format_datetime(value):
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def _current_timestamp_filename(prefix: str, extension: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.{extension}"


def export_transactions_csv(transactions: Iterable[Transaction]) -> HttpResponse:
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f"attachment; filename={_current_timestamp_filename('transactions', 'csv')}"
    )

    writer = csv.writer(response)
    writer.writerow(TRANSACTION_HEADERS)

    for transaction in transactions:
        user = transaction.user
        writer.writerow(
            [
                transaction.id,
                user.id if user else "",
                getattr(user, "first_name", ""),
                getattr(user, "last_name", ""),
                getattr(user, "email", ""),
                getattr(user, "phone_number", ""),
                getattr(user, "national_code", ""),
                getattr(user, "gender", ""),
                getattr(user, "city", ""),
                getattr(user, "university", ""),
                getattr(user, "major", ""),
                transaction.amount or "",
                (transaction.amount or 0) // 10,
                transaction.status,
                _flatten_items(transaction),
                transaction.discount.code if transaction.discount else "",
                transaction.track_id,
                transaction.order_id,
                transaction.ref_number or "",
                transaction.card_number or "",
                transaction.result or "",
                transaction.gateway_response or "",
                _format_datetime(transaction.created_at),
                _format_datetime(transaction.completed_at),
                _format_datetime(transaction.updated_at),
            ]
        )

    return response


def _transactions_to_rows(transactions: Iterable[Transaction]):
    for transaction in transactions:
        user = transaction.user
        yield [
            transaction.id,
            user.id if user else "",
            getattr(user, "first_name", ""),
            getattr(user, "last_name", ""),
            getattr(user, "email", ""),
            getattr(user, "phone_number", ""),
            getattr(user, "national_code", ""),
            getattr(user, "gender", ""),
            getattr(user, "city", ""),
            getattr(user, "university", ""),
            getattr(user, "major", ""),
            transaction.amount or 0,
            (transaction.amount or 0) // 10,
            transaction.status,
            _flatten_items(transaction),
            transaction.discount.code if transaction.discount else "",
            transaction.track_id,
            transaction.order_id,
            transaction.ref_number or "",
            transaction.card_number or "",
            transaction.result or "",
            transaction.gateway_response or "",
            _format_datetime(transaction.created_at),
            _format_datetime(transaction.completed_at),
            _format_datetime(transaction.updated_at),
        ]


def export_transactions_excel(transactions: Iterable[Transaction]) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(TRANSACTION_HEADERS)

    for row in _transactions_to_rows(transactions):
        ws.append(row)

    _auto_fit_columns(ws)

    output = BytesIO()
    wb.save(output)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename={_current_timestamp_filename('transactions', 'xlsx')}"
    )
    return response


def _auto_fit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value else 0
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass
        adjusted_width = max_length + 2 if max_length < 60 else 60
        ws.column_dimensions[column].width = adjusted_width


def export_purchasing_items_excel() -> HttpResponse:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    purchasing_items = list(PurchasingItem.objects.all())
    transactions = (
        Transaction.objects.filter(status="completed")
        .select_related("user", "discount")
        .order_by("-completed_at")
    )

    rows_by_item = {item.name: [] for item in purchasing_items}

    for transaction in transactions:
        user = transaction.user
        try:
            items = json.loads(transaction.items or "[]")
            if not isinstance(items, list):
                continue
        except (TypeError, json.JSONDecodeError):
            continue

        for item_name in items:
            if item_name not in rows_by_item:
                continue

            rows_by_item[item_name].append(
                [
                    f"{getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')}".strip(),
                    getattr(user, "email", ""),
                    getattr(user, "phone_number", ""),
                    getattr(user, "national_code", ""),
                    getattr(user, "gender", ""),
                    getattr(user, "city", ""),
                    getattr(user, "university", ""),
                    getattr(user, "major", ""),
                    transaction.ref_number or transaction.track_id,
                    transaction.track_id,
                    transaction.order_id,
                    transaction.status,
                    (transaction.amount or 0) // 10,
                    _format_datetime(transaction.completed_at),
                ]
            )

    if not purchasing_items:
        ws = wb.create_sheet(title="Purchases")
        ws.append(["No purchasing items defined"])
    else:
        for item in purchasing_items:
            sheet_name = _sanitize_sheet_name(item.name)
            ws = wb.create_sheet(title=sheet_name)
            ws.append(PURCHASING_ITEM_HEADERS)
            for row in rows_by_item[item.name]:
                ws.append(row)
            _auto_fit_columns(ws)

    output = BytesIO()
    wb.save(output)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename={_current_timestamp_filename('purchasing_items', 'xlsx')}"
    )
    return response


def _sanitize_sheet_name(name: str) -> str:
    invalid_chars = set('[]:*?/\\')
    cleaned = "".join(char for char in name if char not in invalid_chars)
    cleaned = cleaned.strip() or "Item"
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    return cleaned
