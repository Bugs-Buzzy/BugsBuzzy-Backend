import csv
import json
from io import BytesIO

from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import User
from .models import Transaction, PurchasingItem


@override_settings(SECURE_SSL_REDIRECT=False)
class AdminExportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin_user = User.objects.create_superuser(
            email="admin@example.com",
            password="supersecret",
            normalized_email="admin@example.com",
            first_name="Admin",
            last_name="User",
            national_code="1234567890",
            phone_number="09123456789",
            gender="M",
        )

        cls.buyer = User.objects.create_user(
            email="buyer@example.com",
            password="buyerpass",
            normalized_email="buyer@example.com",
            first_name="Buyer",
            last_name="Example",
            national_code="1111111111",
            phone_number="09111111111",
            gender="F",
            city="Tehran",
            university="UT",
            major="CS",
        )

        cls.item_one = PurchasingItem.objects.create(
            name="EarlyBird",
            description="Early bird ticket",
            amount=500000,
            initial_count=100,
            purchased_count=1,
            color="#10b981",
        )

        cls.item_two = PurchasingItem.objects.create(
            name="Workshop",
            description="Workshop access",
            amount=300000,
            initial_count=50,
            purchased_count=0,
            color="#3b82f6",
        )

        cls.transaction = Transaction.objects.create(
            user=cls.buyer,
            amount=500000,
            status="completed",
            items=json.dumps([cls.item_one.name]),
            track_id="TRACK123",
            order_id="ORDER123",
            ref_number=987654321,
            card_number="6037****",
            result=100,
            completed_at=timezone.now(),
        )

    def setUp(self):
        self.client.force_login(self.admin_user)

    def test_transaction_export_csv_contains_transaction(self):
        url = reverse("admin:payments_transaction_export_csv")
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])

        content = response.content.decode("utf-8")
        rows = list(csv.reader(content.splitlines()))

        self.assertGreater(len(rows), 1)
        header = rows[0]
        self.assertIn("Amount (Toman)", header)
        first_data_row = rows[1]
        self.assertEqual(first_data_row[0], str(self.transaction.id))
        self.assertEqual(first_data_row[4], self.buyer.email)

    def test_transaction_export_excel_contains_sheet(self):
        url = reverse("admin:payments_transaction_export_excel")
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )

        wb = load_workbook(filename=BytesIO(response.content))
        self.assertIn("Transactions", wb.sheetnames)
        ws = wb["Transactions"]
        self.assertGreater(ws.max_row, 1)
        self.assertEqual(ws.cell(row=2, column=1).value, self.transaction.id)
        self.assertEqual(ws.cell(row=2, column=5).value, self.buyer.email)

    def test_purchasing_item_export_excel_creates_sheet_per_item(self):
        url = reverse("admin:payments_purchasingitem_export_buyers_excel")
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)

        wb = load_workbook(filename=BytesIO(response.content))
        # Ensure that both items are present even if one has no buyers
        sheet_names = set(wb.sheetnames)
        self.assertIn("EarlyBird", sheet_names)
        self.assertIn("Workshop", sheet_names)

        early_bird_sheet = wb["EarlyBird"]
        self.assertEqual(early_bird_sheet.max_row, 2)  # header + one buyer
        self.assertEqual(early_bird_sheet.cell(row=2, column=2).value, self.buyer.email)
